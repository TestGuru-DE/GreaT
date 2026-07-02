"""
routers/api_generate.py – G.R.E.A.T. REST-API für Testfall-Generierung und Export.
"""
from __future__ import annotations

import io
import csv
import json
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session

from ..db import get_db
from .. import models, schemas
from ..services import (
    load_categories_values,
    generate_cases,
    assignment_from_testcase,
    status_for_assignment,
    load_error_values,
    sort_testcases_error_last,
)
from core.rules.rule_engine import RuleEngine, ForbiddenRule, DependencyRule, CombineRule

router = APIRouter(tags=["API"])


@router.post("/projects/{pid}/generate", response_model=schemas.GenerateResponse)
def generate(pid: int, payload: schemas.GenerateRequest, db: Session = Depends(get_db)):
    if payload.limit is not None and payload.limit <= 0:
        raise HTTPException(status_code=400, detail="limit must be positive")

    catmap = load_categories_values(db, pid)
    if not catmap or any(len(v) == 0 for v in catmap.values()):
        raise HTTPException(status_code=400, detail="Project must have categories and values.")

    cases = generate_cases(catmap, payload.strategy, db=db, project_id=pid, t_strength=payload.t_strength)

    # REQ-3005: Regeln anwenden wenn gewünscht
    if payload.apply_rules:
        db_rules = db.query(models.Rule).filter(models.Rule.project_id == pid).all()
        cat_name_by_id = {c.id: c.name for c in db.query(models.Category).filter(models.Category.project_id == pid).all()}
        engine_rules = []
        for r in db_rules:
            ic = cat_name_by_id.get(r.if_category_id, "")
            tc_name = cat_name_by_id.get(r.then_category_id, "")
            if r.type == "exclude":
                engine_rules.append(ForbiddenRule(ic, r.if_value, tc_name, r.then_value))
            elif r.type == "dependency":
                engine_rules.append(DependencyRule(ic, r.if_value, tc_name, r.then_value))
            elif r.type == "combine":
                allowed = json.loads(r.then_values_json or "[]")
                engine_rules.append(CombineRule(ic, r.if_value, tc_name, allowed))
        cases = RuleEngine(engine_rules).apply(cases)

    # BUG-5: Fehlerwert-Testfälle ans Ende sortieren
    error_values = load_error_values(db, pid)
    cases = sort_testcases_error_last(cases, error_values)

    if payload.limit is not None:
        cases = cases[: payload.limit]

    gen = models.Generation(project_id=pid, strategy=payload.strategy)
    db.add(gen)
    db.flush()

    categories = (
        db.query(models.Category)
        .filter(models.Category.project_id == pid)
        .order_by(models.Category.order_index, models.Category.id)
        .all()
    )
    cat_by_name = {c.name: c.id for c in categories}

    for idx, assignment in enumerate(cases, start=1):
        tc = models.TestCase(generation_id=gen.id, name=f"TC_{idx}")
        db.add(tc)
        db.flush()
        for k, v in assignment.items():
            db.add(models.TestCaseValue(testcase_id=tc.id, category_id=cat_by_name[k], value=v))

    db.commit()
    return schemas.GenerateResponse(generation_id=gen.id, count=len(cases))


@router.get("/generations/{gid}/testcases", response_model=List[schemas.TestCaseOut])
def get_testcases(gid: int, db: Session = Depends(get_db)):
    gen = db.get(models.Generation, gid)
    if gen is None:
        raise HTTPException(status_code=404, detail="Generation not found.")

    categories = (
        db.query(models.Category)
        .filter(models.Category.project_id == gen.project_id)
        .order_by(models.Category.order_index, models.Category.id)
        .all()
    )
    name_by_id = {c.id: c.name for c in categories}

    # REQ-3050: risk_weight-Map laden (value_string -> risk_weight)
    all_values = (
        db.query(models.Value)
        .join(models.Category)
        .filter(models.Category.project_id == gen.project_id)
        .all()
    )
    value_risk_map = {v.value: v.risk_weight for v in all_values}

    testcases = (
        db.query(models.TestCase)
        .filter(models.TestCase.generation_id == gid)
        .order_by(models.TestCase.id)
        .all()
    )

    # REQ-3063: Fehlerwerte laden für Markierung
    error_values = load_error_values(db, gen.project_id)

    out: List[dict] = []
    for tc in testcases:
        vals = (
            db.query(models.TestCaseValue)
            .filter(models.TestCaseValue.testcase_id == tc.id)
            .order_by(models.TestCaseValue.id)
            .all()
        )
        assignments = {name_by_id.get(v.category_id, f"cat#{v.category_id}"): v.value for v in vals}
        has_error = any(v in error_values for v in assignments.values())
        
        # REQ-3050: Risikoabdeckung berechnen (Summe der risk_weight)
        risk_coverage = sum(value_risk_map.get(val, 1) for val in assignments.values())
        
        out.append({
            "name": tc.name,
            "assignments": assignments,
            "_has_error_value": has_error,
            "risk_coverage": float(risk_coverage),
        })

    return out


@router.get("/generations/{gen_id}/export/csv")
def export_generation_csv(
    gen_id: int,
    include_status: bool = Query(False, alias="status"),
    encoding: str = Query("utf-8-sig", description="z. B. utf-8-sig, utf-8, cp1252, iso-8859-1, utf-16le, utf-16be, utf-16"),
    excel: bool = Query(True, description="Wenn True: erste Zeile 'sep=;' für Excel"),
    bom: Optional[bool] = Query(None, description="BOM explizit setzen/entfernen; Standard je nach Encoding"),
    db: Session = Depends(get_db),
):
    gen = db.get(models.Generation, gen_id)
    if not gen:
        raise HTTPException(status_code=404, detail="Generation not found.")

    tcs = (
        db.query(models.TestCase)
        .filter(models.TestCase.generation_id == gen_id)
        .order_by(models.TestCase.id)
        .all()
    )
    cats = (
        db.query(models.Category)
        .filter(models.Category.project_id == gen.project_id)
        .order_by(models.Category.order_index, models.Category.id)
        .all()
    )
    cat_headers: List[str] = [c.name for c in cats]

    out = io.StringIO(newline="")
    writer = csv.writer(out, delimiter=";", lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)

    if excel:
        writer.writerow(["sep=;"])

    headers = cat_headers + ["__TestCaseID", "__GenerationID", "__Strategy"]
    if include_status:
        headers.append("Status")
    writer.writerow(headers)

    for tc in tcs:
        a = assignment_from_testcase(db, tc.id)
        row = [a.get(h, "") for h in cat_headers]
        row += [tc.id, gen.id, gen.strategy]
        if include_status:
            row.append(status_for_assignment(gen.project_id, a, db))
        writer.writerow(row)

    csv_text: str = out.getvalue()

    enc_norm = encoding.lower().strip()
    aliases = {"latin1": "iso-8859-1"}
    enc = aliases.get(enc_norm, enc_norm)

    allowed = {"utf-8", "utf-8-sig", "cp1252", "iso-8859-1", "utf-16", "utf-16le", "utf-16be"}
    if enc not in allowed:
        raise HTTPException(status_code=400, detail=f"Unsupported encoding '{encoding}'. Allowed: {sorted(allowed)}")

    if bom is None:
        bom_default = enc in {"utf-8-sig", "utf-16", "utf-16le", "utf-16be"}
    else:
        bom_default = bool(bom)

    content_charset = enc if enc != "utf-8-sig" else "utf-8"

    if enc == "utf-8-sig":
        data_bytes = csv_text.encode("utf-8-sig")
    elif enc in {"utf-16", "utf-16le", "utf-16be"}:
        data_bytes = csv_text.encode(enc if enc != "utf-16" else "utf-16")
        if bom_default:
            if enc == "utf-16le" and not data_bytes.startswith(b"\xff\xfe"):
                data_bytes = b"\xff\xfe" + data_bytes
            elif enc == "utf-16be" and not data_bytes.startswith(b"\xfe\xff"):
                data_bytes = b"\xfe\xff" + data_bytes
        else:
            if data_bytes.startswith(b"\xff\xfe"):
                data_bytes = data_bytes[2:]
            if data_bytes.startswith(b"\xfe\xff"):
                data_bytes = data_bytes[2:]
    else:
        data_bytes = csv_text.encode(enc, errors="strict")
        if bom_default and enc == "utf-8":
            data_bytes = b"\xef\xbb\xbf" + data_bytes

    filename = f"tanos_generation_{gen.id}.csv"
    resp_headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": f"text/csv; charset={content_charset}",
    }
    return Response(content=data_bytes, headers=resp_headers, media_type=f"text/csv; charset={content_charset}")


# ---------------------------------------------------------------------------
# JSON Export – REQ-1001
# ---------------------------------------------------------------------------

@router.get("/generations/{gen_id}/export/json")
def export_generation_json(gen_id: int, db: Session = Depends(get_db)):
    """Exportiert alle Testfälle einer Generation als strukturiertes JSON.

    REQ-1001: JSON Export mit project/categories/testcases-Hierarchie.
    """
    gen = db.get(models.Generation, gen_id)
    if not gen:
        raise HTTPException(status_code=404, detail="Generation not found.")

    categories = (
        db.query(models.Category)
        .filter(models.Category.project_id == gen.project_id)
        .order_by(models.Category.order_index, models.Category.id)
        .all()
    )
    name_by_id = {c.id: c.name for c in categories}

    testcases_db = (
        db.query(models.TestCase)
        .filter(models.TestCase.generation_id == gen_id)
        .order_by(models.TestCase.id)
        .all()
    )

    testcases_out: List[dict] = []
    for tc in testcases_db:
        vals = (
            db.query(models.TestCaseValue)
            .filter(models.TestCaseValue.testcase_id == tc.id)
            .order_by(models.TestCaseValue.id)
            .all()
        )
        assignments = {
            name_by_id.get(v.category_id, f"cat#{v.category_id}"): v.value
            for v in vals
        }
        testcases_out.append({"name": tc.name, "assignments": assignments})

    payload = {
        "generation_id": gen.id,
        "project_id": gen.project_id,
        "strategy": gen.strategy,
        "testcase_count": len(testcases_out),
        "testcases": testcases_out,
    }

    return Response(
        content=json.dumps(payload, ensure_ascii=False, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="tanos_generation_{gen.id}.json"'},
    )


# ---------------------------------------------------------------------------
# Excel Export – REQ-1002
# ---------------------------------------------------------------------------

@router.get("/generations/{gen_id}/export/xlsx")
def export_generation_xlsx(gen_id: int, db: Session = Depends(get_db)):
    """Exportiert alle Testfälle einer Generation als Excel-Datei (.xlsx).

    REQ-1002: Excel Export mit openpyxl – Kategorien als Spaltenüberschriften.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl nicht installiert.")

    gen = db.get(models.Generation, gen_id)
    if not gen:
        raise HTTPException(status_code=404, detail="Generation not found.")

    categories = (
        db.query(models.Category)
        .filter(models.Category.project_id == gen.project_id)
        .order_by(models.Category.order_index, models.Category.id)
        .all()
    )
    cat_headers: List[str] = [c.name for c in categories]
    name_by_id = {c.id: c.name for c in categories}

    testcases_db = (
        db.query(models.TestCase)
        .filter(models.TestCase.generation_id == gen_id)
        .order_by(models.TestCase.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Generation_{gen.id}"

    # Header-Zeile (fett)
    headers = cat_headers + ["TC-ID", "Generierungs-ID", "Strategie"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Datenzeilen
    for row_idx, tc in enumerate(testcases_db, start=2):
        vals = (
            db.query(models.TestCaseValue)
            .filter(models.TestCaseValue.testcase_id == tc.id)
            .order_by(models.TestCaseValue.id)
            .all()
        )
        assignments = {
            name_by_id.get(v.category_id, f"cat#{v.category_id}"): v.value
            for v in vals
        }
        row_data = [assignments.get(h, "") for h in cat_headers]
        row_data += [tc.id, gen.id, gen.strategy]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # In Bytes-Buffer schreiben
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"tanos_generation_{gen.id}.xlsx"
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return Response(
        content=buffer.read(),
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
