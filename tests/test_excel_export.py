# REQ-1002: Excel Export (.xlsx)
# TDD: Tests werden VOR der Implementierung geschrieben.
import uuid
from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)


def _create_project_with_testcases() -> tuple[int, int]:
    """Hilfsfunktion: Projekt mit Kategorien + Generierung anlegen. Gibt (pid, gen_id) zurück."""
    name = f"XLSXExport-{uuid.uuid4().hex[:6]}"
    r = client.post("/api/projects", json={"name": name})
    assert r.status_code == 200
    pid = r.json()["id"]

    r = client.post(f"/api/projects/{pid}/categories", json={"name": "Status", "order_index": 0})
    assert r.status_code == 200
    cid = r.json()["id"]
    client.post(f"/api/categories/{cid}/values", json={"value": "Aktiv"})
    client.post(f"/api/categories/{cid}/values", json={"value": "Inaktiv"})

    r = client.post(f"/api/projects/{pid}/generate", json={"strategy": "each"})
    assert r.status_code == 200
    gen_id = r.json()["generation_id"]
    return pid, gen_id


class TestExcelExport:
    """Excel Export – REQ-1002."""

    def test_export_xlsx_gibt_200(self):
        # REQ-1002: Endpunkt erreichbar
        _, gen_id = _create_project_with_testcases()
        r = client.get(f"/api/generations/{gen_id}/export/xlsx")
        assert r.status_code == 200

    def test_export_xlsx_content_type(self):
        # REQ-1002: Content-Type muss xlsx-MIME-Typ sein
        _, gen_id = _create_project_with_testcases()
        r = client.get(f"/api/generations/{gen_id}/export/xlsx")
        assert "spreadsheetml" in r.headers["content-type"] or "octet-stream" in r.headers["content-type"]

    def test_export_xlsx_content_disposition(self):
        # REQ-1002: Dateiname im Content-Disposition-Header
        _, gen_id = _create_project_with_testcases()
        r = client.get(f"/api/generations/{gen_id}/export/xlsx")
        assert "attachment" in r.headers.get("content-disposition", "")
        assert ".xlsx" in r.headers.get("content-disposition", "")

    def test_export_xlsx_ist_valide_xlsx_datei(self):
        # REQ-1002: Rückgabe ist eine valide .xlsx-Datei (openpyxl lesbar)
        import io
        import openpyxl
        _, gen_id = _create_project_with_testcases()
        r = client.get(f"/api/generations/{gen_id}/export/xlsx")
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert len(wb.sheetnames) >= 1

    def test_export_xlsx_enthaelt_kategorienamen(self):
        # REQ-1002: Erster Worksheet enthält Kategorienamen als Spaltenüberschriften
        import io
        import openpyxl
        _, gen_id = _create_project_with_testcases()
        r = client.get(f"/api/generations/{gen_id}/export/xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Status" in headers

    def test_export_xlsx_nicht_existierende_generation(self):
        # REQ-1002: 404 bei unbekannter Generation
        r = client.get("/api/generations/999999/export/xlsx")
        assert r.status_code == 404
